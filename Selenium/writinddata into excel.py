import openpyxl

# writing same data
# file--->workbook-->Sheets-->Rows-->Cells
# file="C:/Users/sudha/Desktop/test.xlsx"
# workbook=openpyxl.load_workbook(file)
# #sheet=workbook["Data"] # if you have multiple sheets we can use this
# sheet=workbook.active # if you have only one sheet, then you can use this
# #designing excel shhet
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Sudha"
# workbook.save(file)


# multiple data
file="C:/Users/sudha/Desktop/test1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Data"] # if you have multiple sheets we can use this
sheet=workbook.active # if you have only one sheet, then you can use this

##designing excel sheet
sheet.cell(1,1).value="Sudha"
sheet.cell(1,2).value=101

sheet.cell(2,1).value="Arjun"
sheet.cell(2,2).value=102

sheet.cell(3,1).value="Akki"
sheet.cell(3,2).value=103

sheet.cell(4,1).value="Raju"
sheet.cell(4,2).value=104

workbook.save(file)