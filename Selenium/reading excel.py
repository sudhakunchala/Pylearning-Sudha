import openpyxl

# file--->workbook-->Sheets-->Rows-->Cells
file="C:/Users/sudha/Desktop/ReadigExcel.xlsx" # we need specify the path of the excel file
workbook=openpyxl.load_workbook(file)  # loading work book
sheet=workbook["Sheet1"] # need to specify sheet name

rows=sheet.max_row # this will give count of rows
cols=sheet.max_column # count of columns

# Reading all the rows and columns from the excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='       ')
    print()